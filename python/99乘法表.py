import openpyxl

wb=openpyxl.Workbook()
ws=wb.active


ws.title="99乘法表"

ws.sheet_properties.tabColor="f05654"
row_color=["f05654","ff2121","dc3023","ff3300","cb3a56","a98175","b36d61","ef7a82","ff0097"]

for row in range(1,10):
    for column in range(1,row+1):
        cell=ws.cell(row=row,column=column)
        cell_value=str(column)+"×"+str(row)+"="+str(column*row)
        cell.value=cell_value

        font_set=openpyxl.styles.Font(name="Arial",size=14,italic=True,color="000000",bold=True)
        cell.font=font_set

        border=openpyxl.styles.Border(top=openpyxl.styles.Side(border_style="thin",color="FF000000"),bottom=openpyxl.styles.Side(border_style="thin",color="FF000000"),left=openpyxl.styles.Side(border_style="thin",color="FF000000"),right=openpyxl.styles.Side(border_style="thin",color="FF000000"))
        cell.border=border

        fill=openpyxl.styles.PatternFill("solid",fgColor=row_color[row-1])
        cell.fill=fill

        

wb.save("九九乘法表.xlsx")
print("表格创建成功！")
